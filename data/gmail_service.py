import base64
import os
import re
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from PyPDF2 import PdfReader
from google.auth.exceptions import RefreshError
from google.auth.transport.requests import AuthorizedSession, Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from openpyxl import load_workbook

from data.gmail_store import init_db, replace_attachments, upsert_message


SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]
ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CREDENTIALS_PATH = ROOT / "credentials" / "gmail_credentials.json"
DEFAULT_TOKEN_PATH = ROOT / "credentials" / "gmail_token.json"
DEFAULT_ATTACHMENT_DIR = ROOT / "data" / "gmail_attachments"
GMAIL_API_ROOT = "https://gmail.googleapis.com/gmail/v1/users/me"
DEFAULT_LOOKBACK_DAYS = 7
MAX_LOOKBACK_DAYS = 90


class GmailShippingDataService:
    def __init__(self):
        self.credentials_path = Path(
            os.getenv("GMAIL_OAUTH_CREDENTIALS", DEFAULT_CREDENTIALS_PATH)
        )
        self.token_path = Path(os.getenv("GMAIL_OAUTH_TOKEN", DEFAULT_TOKEN_PATH))
        self.attachment_dir = Path(
            os.getenv("GMAIL_ATTACHMENT_DIR", DEFAULT_ATTACHMENT_DIR)
        )
        self.lookback_days = self._read_lookback_days()

    def ensure_oauth_token(self):
        creds = self._load_credentials(interactive=True)
        self.token_path.parent.mkdir(parents=True, exist_ok=True)
        self.token_path.write_text(creds.to_json(), encoding="utf-8")
        return str(self.token_path)

    def sync_recent_shipping_data(self):
        init_db()
        session = self._authorized_session()
        label_id = self._find_label_id(session, "shipping-data")
        if not label_id:
            raise RuntimeError("Gmail label 'shipping-data' not found.")

        queries = [
            "newer_than:1d",
            f'subject:"SSY SINGAPORE" newer_than:{self.lookback_days}d',
        ]
        if self.lookback_days > 1:
            queries.append(f"newer_than:{self.lookback_days}d")
        message_refs = []
        seen_ids = set()
        for query in queries:
            response = self._get_json(
                session,
                f"{GMAIL_API_ROOT}/messages?labelIds={quote(label_id)}&q={quote(query)}&maxResults=100",
            )
            for ref in response.get("messages", []):
                message_id = ref.get("id")
                if not message_id or message_id in seen_ids:
                    continue
                seen_ids.add(message_id)
                message_refs.append(ref)

        synced = []
        for ref in message_refs:
            message = self._get_json(
                session,
                f"{GMAIL_API_ROOT}/messages/{ref['id']}?format=full",
            )
            synced.append(self._save_message(session, message))

        return {
            "synced_count": len(synced),
            "message_ids": synced,
            "label": "shipping-data",
            "query": f"newer_than:1d + SSY SINGAPORE within {self.lookback_days}d",
            "synced_at": datetime.now().isoformat(),
        }

    def _save_message(self, session, message):
        payload = message.get("payload", {})
        headers = self._headers_to_map(payload.get("headers", []))
        body_text = self._extract_plain_text(payload)
        body_summary = self._summarize_text(
            body_text or message.get("snippet") or "", limit=220
        )
        attachments = self._collect_attachments(session, payload, message["id"])

        record = {
            "gmail_message_id": message["id"],
            "thread_id": message.get("threadId"),
            "label_ids": message.get("labelIds", []),
            "sender": headers.get("From", ""),
            "subject": headers.get("Subject", ""),
            "internal_ts": int(message.get("internalDate") or 0),
            "received_at": self._format_ts(message.get("internalDate")),
            "snippet": message.get("snippet", ""),
            "body_text": body_text,
            "body_summary": body_summary,
            "has_attachments": bool(attachments),
            "raw_payload_json": payload,
            "synced_at": datetime.now().isoformat(),
        }
        upsert_message(record)
        replace_attachments(message["id"], attachments)
        return message["id"]

    def _load_credentials(self, interactive):
        creds = None
        if self.token_path.exists():
            creds = Credentials.from_authorized_user_file(str(self.token_path), SCOPES)
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
                self._save_credentials(creds)
            except RefreshError as exc:
                raise RuntimeError(
                    "Gmail token 已过期或被撤销。请重新运行 "
                    "`py scripts\\gmail_oauth_setup.py` 完成授权；"
                    "如果是 GitHub Actions 报错，还需要同步更新仓库 Secret "
                    "`GMAIL_TOKEN_JSON`。"
                ) from exc
        if creds and creds.valid:
            return creds
        if not self.credentials_path.exists():
            raise RuntimeError(
                f"未找到 Gmail OAuth 客户端文件: {self.credentials_path}"
            )
        if not interactive:
            raise RuntimeError(
                "Gmail OAuth 尚未完成。请先运行 `py scripts\\gmail_oauth_setup.py` 完成授权。"
            )
        flow = InstalledAppFlow.from_client_secrets_file(
            str(self.credentials_path), SCOPES
        )
        return flow.run_local_server(port=0)

    def _save_credentials(self, creds):
        self.token_path.parent.mkdir(parents=True, exist_ok=True)
        self.token_path.write_text(creds.to_json(), encoding="utf-8")

    @staticmethod
    def _read_lookback_days():
        raw_value = os.getenv("GMAIL_LOOKBACK_DAYS", str(DEFAULT_LOOKBACK_DAYS))
        try:
            days = int(raw_value)
        except (TypeError, ValueError):
            days = DEFAULT_LOOKBACK_DAYS
        return max(1, min(days, MAX_LOOKBACK_DAYS))

    def _authorized_session(self):
        creds = self._load_credentials(interactive=False)
        session = AuthorizedSession(creds)
        session.timeout = 30
        return session

    def _find_label_id(self, session, label_name):
        response = self._get_json(session, f"{GMAIL_API_ROOT}/labels")
        for item in response.get("labels", []):
            if item.get("name", "").lower() == label_name.lower():
                return item.get("id")
        return None

    @staticmethod
    def _headers_to_map(headers):
        return {item.get("name"): item.get("value") for item in headers}

    def _extract_plain_text(self, payload):
        parts = list(self._walk_parts(payload))
        for part in parts:
            if part.get("mimeType") == "text/plain":
                data = part.get("body", {}).get("data")
                if data:
                    return self._normalize_text_blocks(self._decode_base64(data))
        for part in parts:
            if part.get("mimeType") == "text/html":
                data = part.get("body", {}).get("data")
                if data:
                    return self._html_to_text(self._decode_base64(data))
        body_data = payload.get("body", {}).get("data")
        if body_data:
            return self._normalize_text_blocks(self._decode_base64(body_data))
        return ""

    def _collect_attachments(self, session, payload, message_id):
        attachments = []
        now = datetime.now().isoformat()
        for part in self._walk_parts(payload):
            filename = part.get("filename") or ""
            body = part.get("body", {})
            attachment_id = body.get("attachmentId")
            mime_type = part.get("mimeType")
            if not filename or not attachment_id:
                continue
            if not self._is_supported_attachment(filename, mime_type):
                continue
            blob = self._get_json(
                session,
                f"{GMAIL_API_ROOT}/messages/{message_id}/attachments/{attachment_id}",
            )
            data = self._decode_base64_bytes(blob.get("data", ""))
            safe_name = self._safe_filename(filename)
            target_dir = self.attachment_dir / datetime.now().strftime("%Y%m%d")
            target_dir.mkdir(parents=True, exist_ok=True)
            local_path = target_dir / f"{message_id}_{safe_name}"
            local_path.write_bytes(data)
            parsed_text = self._parse_attachment(local_path, mime_type)
            attachments.append(
                {
                    "attachment_id": attachment_id,
                    "filename": filename,
                    "mime_type": mime_type,
                    "size_bytes": len(data),
                    "local_path": str(local_path),
                    "parsed_text": parsed_text,
                    "parsed_summary": self._summarize_text(parsed_text, limit=320),
                    "created_at": now,
                }
            )
        return attachments

    def _parse_attachment(self, path, mime_type):
        suffix = path.suffix.lower()
        try:
            if suffix == ".pdf" or mime_type == "application/pdf":
                reader = PdfReader(str(path))
                texts = []
                for page in reader.pages[:20]:
                    texts.append(page.extract_text() or "")
                return self._normalize_text_blocks("\n\n".join(texts))
            if (
                suffix in {".xlsx", ".xls"}
                or "sheet" in (mime_type or "").lower()
                or "excel" in (mime_type or "").lower()
            ):
                workbook = load_workbook(
                    filename=str(path), read_only=True, data_only=True
                )
                chunks = []
                for sheet in workbook.worksheets[:5]:
                    chunks.append(f"[{sheet.title}]")
                    rows = []
                    for row in sheet.iter_rows(
                        min_row=1, max_row=20, values_only=True
                    ):
                        values = ["" if cell is None else str(cell) for cell in row]
                        if any(values):
                            rows.append(" | ".join(values))
                    chunks.append("\n".join(rows))
                return self._normalize_text_blocks("\n\n".join(chunks))
        except Exception as exc:
            return f"解析失败: {exc}"
        return "暂不支持该附件类型的解析。"

    @staticmethod
    def _walk_parts(payload):
        parts = payload.get("parts") or []
        if not parts:
            yield payload
            return
        for part in parts:
            yield part
            if part.get("parts"):
                yield from GmailShippingDataService._walk_parts(part)

    @staticmethod
    def _decode_base64(value):
        return GmailShippingDataService._decode_base64_bytes(value).decode(
            "utf-8", errors="ignore"
        )

    @staticmethod
    def _decode_base64_bytes(value):
        padding = "=" * (-len(value) % 4)
        return base64.urlsafe_b64decode((value + padding).encode("utf-8"))

    @staticmethod
    def _html_to_text(html):
        html = re.sub(r"<\s*br\s*/?\s*>", "\n", html, flags=re.IGNORECASE)
        html = re.sub(
            r"</\s*(p|div|li|tr|h[1-6])\s*>", "\n", html, flags=re.IGNORECASE
        )
        text = re.sub(r"<[^>]+>", " ", html)
        return GmailShippingDataService._normalize_text_blocks(text)

    @staticmethod
    def _normalize_whitespace(text):
        return re.sub(r"\s+", " ", (text or "")).strip()

    @staticmethod
    def _normalize_text_blocks(text):
        if not text:
            return ""
        text = text.replace("\r\n", "\n").replace("\r", "\n")
        lines = [re.sub(r"[ \t]+", " ", line).strip() for line in text.split("\n")]
        merged = []
        blank = False
        for line in lines:
            if line:
                merged.append(line)
                blank = False
            elif not blank:
                merged.append("")
                blank = True
        return "\n".join(merged).strip()

    @staticmethod
    def _format_ts(internal_date):
        if not internal_date:
            return None
        return datetime.fromtimestamp(int(internal_date) / 1000).isoformat()

    @staticmethod
    def _summarize_text(text, limit=220):
        text = GmailShippingDataService._normalize_whitespace(text)
        if not text:
            return ""
        return text[:limit] + ("..." if len(text) > limit else "")

    @staticmethod
    def _safe_filename(name):
        safe = re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("._")
        if not safe:
            return "attachment"
        max_length = 120
        if len(safe) <= max_length:
            return safe
        suffix = Path(safe).suffix
        stem = Path(safe).stem if suffix else safe
        max_stem_length = max(1, max_length - len(suffix))
        return f"{stem[:max_stem_length].rstrip('._-')}{suffix}"

    @staticmethod
    def _is_supported_attachment(filename, mime_type):
        suffix = Path(filename).suffix.lower()
        if suffix in {".pdf", ".xlsx", ".xls"}:
            return True
        mime = (mime_type or "").lower()
        return mime == "application/pdf" or "excel" in mime or "sheet" in mime

    @staticmethod
    def _get_json(session, url):
        response = session.get(url, timeout=30)
        if response.status_code >= 400:
            raise RuntimeError(
                f"Gmail API 请求失败: {response.status_code} {response.text[:300]}"
            )
        return response.json()
